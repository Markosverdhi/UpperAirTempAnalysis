from openpyxl import load_workbook
import numpy as np
from matplotlib import pyplot as plt, dates as mdates, ticker as tick
import pandas as pd
import tkinter as tk
from tkinter import *
from tkinter import ttk, Button
import statistics as s

#Developed by Justin Ngo, Markos Verdhi

#read spreadsheets and workbook
wb = load_workbook('Upper_Air_temps_data.xlsx')
tropoArr = []
stratoArr = []
troposphere = wb["Sheet1"]
stratosphere = wb["Sheet2"]

#array for tropospheric data
for col in troposphere.iter_rows(min_row=2, max_col=7, max_row=519, values_only=True):
    tropoArr.append(col)

#array for stratospheric data
for col in stratosphere.iter_rows(min_row=2, max_col=7, max_row=519, values_only=True):
    stratoArr.append(col)

#modified binary search method for openpyxl
def findInterval(arr, startMonth, startYear, endMonth, endYear, r=None, l=0):
    returns = []
    if r == None:
        r=len(arr)
    while len(returns)==0:
        middle = l+(r-l)//2
        if arr[middle][0] == startYear:
            if startMonth == l:
                returns.append(middle + 2)
            elif startMonth != l:
                diff = startMonth - arr[middle][1]
                startIndex = (middle + diff)
                returns.append(startIndex)
            monthdiff = endMonth - startMonth
            yeardiff = endYear - startYear
            endIndex = returns[0] + (12*yeardiff + monthdiff)
            if endIndex == len(arr):
                returns.append(endIndex-2)
            else:
                returns.append(endIndex)
            if returns[0] < 0 or returns[1] < 0:
                print('one or more entries in specified range do not exist in the data.')
            else:
                return returns

        elif arr[middle][0] > startYear:
            return findInterval(arr, startMonth, startYear, endMonth, endYear, middle, l)
        elif arr[middle][0] < startYear:
            return findInterval(arr, startMonth, startYear, endMonth, endYear, r, middle)
        else:
            print('one or more entries in specified range do not exist in the data.')
            return None

# method for graphing data
def graphBySeason(arr, startMonth, startYear, endMonth, endYear, hemi):
    #searches for start and end index
    i = findInterval(arr, startMonth, startYear, endMonth, endYear)
    #create the space for the plots
    fig, ax = plt.subplots(2,2)
    if i[1]-i[0] > 180:
        fig.set_figwidth(12)
        fig.set_figheight(7)
    else:
        fig.set_figwidth(9)
        fig.set_figheight(9)

    #creating cleaned lists for plotting
    xsummer = []
    ysummer = []
    xwinter = []
    ywinter = []
    xspring = []
    yspring = []
    xfall = []
    yfall = []
    for entry in range(i[0],i[1]):
        if arr[entry][1] in {12, 1, 2}:
            ywinter.append(arr[entry][hemi])
            xwinter.append(mdates.datestr2num(f'{arr[entry][1]}/{arr[entry][0]}'))
        elif arr[entry][1] in {3, 4, 5}:
            yspring.append(arr[entry][hemi])
            xspring.append(mdates.datestr2num(f'{arr[entry][1]}/{arr[entry][0]}'))
        elif arr[entry][1] in {6,7,8}:
            ysummer.append(arr[entry][hemi])
            xsummer.append(mdates.datestr2num(f'{arr[entry][1]}/{arr[entry][0]}'))
        elif arr[entry][1] in {9, 10, 11}:
            yfall.append(arr[entry][hemi])
            xfall.append(mdates.datestr2num(f'{arr[entry][1]}/{arr[entry][0]}'))

    #creating plots
    ax[0][0].scatter(xwinter, ywinter, color = 'lightskyblue')
    ax[0][0].set_title('Winter', x=0.9, y=0.9)
    ax[0][1].scatter(xspring, yspring, color = 'palegreen')
    ax[0][1].set_title('Spring', x=0.9, y=0.9)
    ax[1][0].scatter(xsummer, ysummer, color = 'gold')
    ax[1][0].set_title('Summer', x=0.9, y=0.9)
    ax[1][1].scatter(xfall, yfall, color = 'chocolate')
    ax[1][1].set_title('Fall', x=0.9, y=0.9)

    for i in range(2):
        for j in range(2):
            ax[i][j].set(ylabel = 'temp anomaly')
            #YYYY formatting
            ax[i][j].xaxis.set_major_formatter(mdates.ConciseDateFormatter(tick.AutoLocator()))
            #set ticks for xaxis
            ax[i][j].xaxis.set_major_locator(mdates.YearLocator())
            ax[i][j].xaxis.set_minor_locator(mdates.MonthLocator())
        #rotation of labels
            if len(xwinter) > 45:
                for label in ax[i][j].get_xticklabels(which = 'major')[::2]:
                    label.set_visible(False)

            for label in ax[i][j].get_xticklabels(which = 'major'):
                label.set(rotation=45, horizontalalignment='right')
    plt.show()

def fiveNumberSum(arr, startMonth, startYear, endMonth, endYear, hemi):
    i = findInterval(arr, startMonth, startYear, endMonth, endYear)
    cData = []
    for entry in range(i[0],i[1]):
        cData.append(arr[entry][hemi])

    minimum = min(cData)
    maximum = max(cData)
    returns = [minimum, maximum]
    for i in range(3) : returns.insert(-1,float(format(s.quantiles(cData, n=4)[i-1], '.3f')))
    return returns

def boxPlot(arr, startMonth, startYear, endMonth, endYear, hemi):
    fns = fiveNumberSum(arr, startMonth, startYear, endMonth, endYear, hemi)
    box = plt.figure(figsize = (10, 7))
    ax = box.add_axes([0.1,0.1,0.8,0.8])
    plt.yticks(color='white')
    ax.boxplot(fns, vert=False)
    plt.show()

#Tkinter

#init the GUI
root = tk.Tk()
root.title("Fuckin Retard")
root.geometry( "1050x200" )

#define dropdown menu values
Months = ["select Month", "January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"]

Years = ["Select Year", 1978, 1979,
              1980, 1981, 1982, 1983, 1984, 1985, 1986, 1987, 1988, 1989,
              1990, 1991, 1992, 1993, 1994, 1995, 1996, 1997, 1998, 1999,
              2000, 2001, 2002, 2003, 2004, 2005, 2006, 2007, 2008, 2009,
              2000, 2001, 2002, 2003, 2004, 2005, 2006, 2007, 2008, 2009,
              2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019,
              2020, 2021]

anomaly_Type = ["Select Anomaly", "Global Average", "Northern Hemisphere",
                "Southern Hemisphere", "Global Land", "Global Ocean"]

altitude = ["Select Region", "Troposphere","Stratosphere"]

chart_Type = ["Select Graph Type", "Scatter Plot", "Box and Whisker"]

#method for running code once user inputs variables
def runShit():
    #exception handling
    if sStart_Month.current() == 0 and sStart_Year.current() == 0 and sEnd_Year.current() == 0 and \
            sEnd_Month.current() == 0 and sAnomaly.current() == 0 and sAlti.current() == 0 and sChart.current() == 0:
        print("Wow you must be a special kind of idiot, you actually managed to --somehow-- leave every field blank")
    elif sStart_Month.current() == 0 or sStart_Year.current() == 0 or sEnd_Year.current() == 0 or \
            sEnd_Month.current() == 0 or sAnomaly.current() == 0 or sAlti.current() == 0 or sChart.current() == 0:
        print("You dun goofed up mufuker, you left a field blank")
    else:
        #decision tree --
        #determine which chart user wants
            #determine which array to read
        if sChart.current() == 1:
            if sAlti.current() == 1:
                graphBySeason(tropoArr, sStart_Month.current(), int(sStart_Year.get()), sEnd_Month.current(), int(sEnd_Year.get()), (sAnomaly.current()+1))
            else:
                graphBySeason(stratoArr, sStart_Month.current(), int(sStart_Year.get()), sEnd_Month.current(), int(sEnd_Year.get()), (sAnomaly.current()+1))
        elif sChart.current() == 2:
            if sAlti.current() == 1:
                boxPlot(tropoArr, sStart_Month.current(), int(sStart_Year.get()), sEnd_Month.current(), int(sEnd_Year.get()), (sAnomaly.current() + 1))
            else:
                boxPlot(stratoArr, sStart_Month.current(), int(sStart_Year.get()), sEnd_Month.current(), int(sEnd_Year.get()), (sAnomaly.current() + 1))
        else:
            print("something got fucked, shouldn't have made it this far")

#create a run button
run = tk.Button(root, command = runShit, width = 15, activeforeground= "green", activebackground= "green", text = "RUN")
run.pack()

#make the dropdown menus
sStart_Month = ttk.Combobox(root, value = Months)
sStart_Month.current(0)
sStart_Month.pack(padx=5, pady=10, side=tk.LEFT)

sStart_Year = ttk.Combobox(root,value = Years)
sStart_Year.current(0)
sStart_Year.pack(padx=5, pady=10, side=tk.LEFT)

sEnd_Month = ttk.Combobox(root,value = Months)
sEnd_Month.current(0)
sEnd_Month.pack(padx=5, pady=10, side=tk.LEFT)

sEnd_Year = ttk.Combobox(root,value = Years)
sEnd_Year.current(0)
sEnd_Year.pack(padx=5, pady=10, side=tk.LEFT)

sAnomaly = ttk.Combobox(root,value = anomaly_Type)
sAnomaly.current(0)
sAnomaly.pack(padx=5, pady=10, side=tk.LEFT)

sAlti = ttk.Combobox(root, value = altitude)
sAlti.current(0)
sAlti.pack(padx=5, pady=10, side=tk.LEFT)

sChart = ttk.Combobox(root, value = chart_Type)
sChart.current(0)
sChart.pack(padx=5, pady=10, side=tk.LEFT)

root.mainloop()
